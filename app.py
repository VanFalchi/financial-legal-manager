import os
import secrets
from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from dateutil.relativedelta import relativedelta
import click
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle
from io import BytesIO
from sqlalchemy import or_, func

import calculations

# --- INITIALIZATION ---
db = SQLAlchemy()
login_manager = LoginManager()
app = Flask(__name__)

# --- CONFIGURATION ---
db_url = os.environ.get("DATABASE_URL")
secret_key = os.environ.get("SECRET_KEY")
if not db_url or not secret_key:
    raise RuntimeError("ERROR: DATABASE_URL and SECRET_KEY environment variables must be configured.")

app.config["SECRET_KEY"] = secret_key
app.config["SQLALCHEMY_DATABASE_URI"] = db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db.init_app(app)
login_manager.init_app(app)
login_manager.login_view = 'login'

# --- MODELS ---
@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    def set_password(self, password): self.password_hash = generate_password_hash(password)
    def check_password(self, password): return check_password_hash(self.password_hash, password)

class Client(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(150), nullable=False)
    tax_id = db.Column(db.String(14), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=True)
    phone = db.Column(db.String(20), nullable=True)
    actions = db.relationship('LegalAction', backref='client', lazy=True, cascade="all, delete-orphan")

class LegalAction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    case_number = db.Column(db.String(100), unique=True, nullable=False)
    action_type = db.Column(db.String(100), nullable=False)
    responsible_partner = db.Column(db.String(150), nullable=False)
    responsible_consultant = db.Column(db.String(150), nullable=True)
    client_id = db.Column(db.Integer, db.ForeignKey('client.id'), nullable=False)
    transactions = db.relationship('Transaction', backref='action', lazy=True, cascade="all, delete-orphan")

class Transaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    value_type = db.Column(db.String(50), nullable=False)
    gross_amount = db.Column(db.Float, nullable=False)
    payment_date = db.Column(db.Date, nullable=False)
    notes = db.Column(db.Text, nullable=True)
    action_id = db.Column(db.Integer, db.ForeignKey('legal_action.id'), nullable=False)
    client_share = db.Column(db.Float)
    finance_manager_share = db.Column(db.Float)
    partner_share = db.Column(db.Float)
    consultant_commission = db.Column(db.Float)

# --- ROUTES ---

@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated: return redirect(url_for('clients'))
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form['username']).first()
        if user and user.check_password(request.form['password']):
            login_user(user)
            return redirect(url_for('clients'))
        else:
            flash('Invalid username or password.', 'error')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/clients', methods=['GET', 'POST'])
@login_required
def clients():
    if request.method == 'POST':
        new_client = Client(
            full_name=request.form['full_name'],
            tax_id=request.form['tax_id'],
            email=request.form['email'],
            phone=request.form['phone']
        )
        try:
            db.session.add(new_client)
            db.session.commit()
            flash('Client registered successfully!', 'success')
            return redirect(url_for('clients'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error registering client: {e}', 'error')
    
    query = request.args.get('q')
    if query:
        search_term = f"%{query}%"
        all_clients = Client.query.filter(or_(Client.full_name.ilike(search_term), Client.tax_id.ilike(search_term))).order_by(func.lower(Client.full_name)).all()
    else:
        all_clients = Client.query.order_by(func.lower(Client.full_name)).all()
    return render_template('clients.html', clients=all_clients)

@app.route('/client/<int:client_id>')
@login_required
def client_detail(client_id):
    client = db.session.get(Client, client_id)
    return render_template('client_detail.html', client=client)

@app.route('/client/<int:client_id>/edit', methods=['GET', 'POST'])
@login_required
def client_edit(client_id):
    client = db.session.get(Client, client_id)
    if not client:
        flash('Client not found.', 'error')
        return redirect(url_for('clients'))
    if request.method == 'POST':
        try:
            client.full_name = request.form['full_name']
            client.tax_id = request.form['tax_id']
            client.email = request.form['email']
            client.phone = request.form['phone']
            db.session.commit()
            flash('Client updated successfully!', 'success')
            return redirect(url_for('client_detail', client_id=client.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating client: {e}', 'error')
    return render_template('client_edit.html', client=client)

@app.route('/client/<int:client_id>/delete', methods=['POST'])
@login_required
def client_delete(client_id):
    client = db.session.get(Client, client_id)
    if client:
        try:
            db.session.delete(client)
            db.session.commit()
            flash('Client and all their actions were deleted successfully.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error deleting client: {e}', 'error')
    else:
        flash('Client not found.', 'error')
    return redirect(url_for('clients'))

@app.route('/client/<int:client_id>/new_action', methods=['GET', 'POST'])
@login_required
def new_action(client_id):
    client = db.session.get(Client, client_id)
    if request.method == 'POST':
        new_action_obj = LegalAction(
            case_number=request.form['case_number'],
            action_type=request.form['action_type'],
            responsible_partner=request.form['responsible_partner'],
            responsible_consultant=request.form.get('responsible_consultant'),
            client_id=client.id
        )
        db.session.add(new_action_obj)
        db.session.commit()
        return redirect(url_for('client_detail', client_id=client.id))
    return render_template('legal_action_form.html', client=client)

@app.route('/action/<int:action_id>', methods=['GET', 'POST'])
@login_required
def action_detail(action_id):
    action = db.session.get(LegalAction, action_id)
    if not action:
        return redirect(url_for('clients'))
    if request.method == 'POST':
        gross_amount = float(request.form['gross_amount'])
        value_type = request.form['value_type']
        has_consultant = True if action.responsible_consultant else False
        shares = calculations.calculate_division(action.action_type, gross_amount, value_type, has_consultant)
        new_transaction = Transaction(
            value_type=value_type,
            gross_amount=gross_amount,
            payment_date=datetime.strptime(request.form['payment_date'], '%Y-%m-%d').date(),
            notes=request.form['notes'],
            action_id=action.id,
            client_share=shares['client'],
            finance_manager_share=shares['FinanceManager'],
            partner_share=shares['partner'],
            consultant_commission=shares['consultant']
        )
        db.session.add(new_transaction)
        db.session.commit()
        flash('Single payment recorded successfully!', 'success')
        return redirect(url_for('action_detail', action_id=action.id))
    return render_template('action_detail.html', action=action)

@app.route('/action/<int:action_id>/edit', methods=['GET', 'POST'])
@login_required
def action_edit(action_id):
    action = db.session.get(LegalAction, action_id)
    if not action:
        flash('Action not found.', 'error')
        return redirect(url_for('clients'))
    if request.method == 'POST':
        try:
            action.case_number = request.form['case_number']
            action.action_type = request.form['action_type']
            action.responsible_partner = request.form['responsible_partner']
            action.responsible_consultant = request.form.get('responsible_consultant')
            db.session.commit()
            flash('Action updated successfully!', 'success')
            return redirect(url_for('action_detail', action_id=action.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating action: {e}', 'error')
    return render_template('action_edit.html', action=action)

@app.route('/action/<int:action_id>/delete', methods=['POST'])
@login_required
def action_delete(action_id):
    action = db.session.get(LegalAction, action_id)
    if action:
        client_id = action.client.id
        try:
            db.session.delete(action)
            db.session.commit()
            flash('Action deleted successfully.', 'success')
            return redirect(url_for('client_detail', client_id=client_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error deleting action: {e}', 'error')
    else:
        flash('Action not found.', 'error')
    return redirect(url_for('clients'))

@app.route('/transaction/<int:transaction_id>/delete', methods=['POST'])
@login_required
def transaction_delete(transaction_id):
    transaction = db.session.get(Transaction, transaction_id)
    if transaction:
        action_id = transaction.action.id
        try:
            db.session.delete(transaction)
            db.session.commit()
            flash('Transaction deleted successfully.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error deleting transaction: {e}', 'error')
        return redirect(url_for('action_detail', action_id=action_id))
    else:
        flash('Transaction not found.', 'error')
        return redirect(url_for('clients'))

@app.route('/action/<int:action_id>/add_installments', methods=['GET', 'POST'])
@login_required
def add_installments(action_id):
    action = db.session.get(LegalAction, action_id)
    if not action:
        return redirect(url_for('clients'))
    if request.method == 'POST':
        installment_amount = float(request.form['installment_amount'])
        number_of_installments = int(request.form['number_of_installments'])
        first_installment_date = datetime.strptime(request.form['first_installment_date'], '%Y-%m-%d').date()
        has_consultant = True if action.responsible_consultant else False
        for i in range(number_of_installments):
            current_installment_date = first_installment_date + relativedelta(months=i)
            shares = calculations.calculate_division(action.action_type, installment_amount, "Monthly Installment", has_consultant)
            new_transaction = Transaction(
                value_type="Monthly Installment",
                gross_amount=installment_amount,
                payment_date=current_installment_date,
                notes=f"Installment {i+1} of {number_of_installments}",
                action_id=action.id,
                client_share=shares['client'],
                finance_manager_share=shares['FinanceManager'],
                partner_share=shares['partner'],
                consultant_commission=shares['consultant']
            )
            db.session.add(new_transaction)
        db.session.commit()
        flash(f'{number_of_installments} installments were added successfully!', 'success')
        return redirect(url_for('action_detail', action_id=action.id))
    return render_template('add_installments.html', action=action)

@app.route('/export_report', methods=['POST'])
@login_required
def export_report():
    start_date_str = request.form['start_date']
    end_date_str = request.form['end_date']
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    transactions = Transaction.query.filter(Transaction.payment_date >= start_date, Transaction.payment_date <= end_date).order_by(Transaction.payment_date).all()
    if not transactions:
        flash('No transactions found in the selected period.', 'info')
        return redirect(url_for('clients'))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Financial Report"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
    currency_style = NamedStyle(name='currency', number_format='"$" #,##0.00')
    headers = ["Payment Date", "Client", "Tax ID", "Case Number", "Action Type", "Partner", "Consultant", "Value Type", "Gross Amount", "Client Share", "FinanceManager Share", "Partner Share", "Consultant Commission", "Notes"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    for transaction in transactions:
        row_data = [transaction.payment_date, transaction.action.client.full_name, transaction.action.client.tax_id, transaction.action.case_number, transaction.action.action_type, transaction.action.responsible_partner, transaction.action.responsible_consultant or "N/A", transaction.value_type, transaction.gross_amount, transaction.client_share, transaction.finance_manager_share, transaction.partner_share, transaction.consultant_commission, transaction.notes]
        sheet.append(row_data)
        current_row = sheet.max_row
        sheet.cell(row=current_row, column=1).number_format = 'DD/MM/YYYY'
        for col_idx in range(9, 14):
             sheet.cell(row=current_row, column=col_idx).style = currency_style
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f'financial_report_{start_date_str}_to_{end_date_str}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# --- CLI COMMANDS ---
@app.cli.command("create-initial-users")
def create_initial_users():
    """Creates initial users from environment variables."""
    for i in range(1, 4):
        user_details_str = os.environ.get(f'USER_{i}_DETAILS')
        if user_details_str:
            try:
                username, password = user_details_str.split(',')
                if not User.query.filter_by(username=username).first():
                    new_user = User(username=username)
                    new_user.set_password(password)
                    db.session.add(new_user)
                    print(f"User '{username}' created.")
                else:
                    print(f"User '{username}' already exists.")
            except ValueError:
                print(f"ERROR: Variable USER_{i}_DETAILS is malformed. Use 'username,password'.")
    db.session.commit()
    print("Initial user verification completed.")

# --- UPDATED PASSWORD RESET COMMAND ---
@app.cli.command("reset-password")
@click.argument("username")
@click.argument("new_password")
def reset_password(username, new_password):
    """Resets a user's password to a specific password."""
    user = User.query.filter_by(username=username).first()
    if not user:
        print(f"Error: User '{username}' not found.")
        return
    
    user.set_password(new_password)
    db.session.commit()
    
    print(f"Password for user '{username}' has been reset successfully!")

# --- INITIAL DATABASE CREATION ---
with app.app_context():
    db.create_all()
